# -*- coding: utf-8 -*-
"""
ThamDinhDuToanApp - Máy chủ Flask Server
Phục vụ quản lý cơ sở giá và luồng trao đổi KHVT vs TTĐ
"""
import os
import json
import re
from datetime import datetime
from flask import Flask, render_template, request, jsonify, send_file
from flask_cors import CORS
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import imis_core
import quote_matcher
import msc_matcher

app = Flask(__name__, static_folder='frontend/dist', static_url_path='')
app.config['JSON_AS_ASCII'] = False
CORS(app)

DATA_DIR = os.path.join(os.path.abspath(os.path.dirname(__file__)), "data")
PROJECTS_DIR = os.path.join(DATA_DIR, "projects")
os.makedirs(PROJECTS_DIR, exist_ok=True)
DATA_FILE = os.path.join(DATA_DIR, "current_dossier.json")
ACTIVE_PROJECT_FILE = os.path.join(DATA_DIR, "active_project.json")


def load_dossier_data():
    if os.path.exists(DATA_FILE):
        try:
            with open(DATA_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
            
    # Dữ liệu mẫu ban đầu mô phỏng Gói 308
    return {
        "dossier_name": "Gói 308 - Mua sắm vật tư SCTX đợt 8 năm 2026",
        "creator": "Nguyễn Anh Hiếu",
        "department": "Tổ Thẩm định Dự toán - NMNĐ Vĩnh Tân 4",
        "items": [
            {
                "id": 1,
                "ma_vt": "3.82.63.134.ENG.00.000",
                "part_no": "IUX 760 MI",
                "ten_vt": "Module đầu vào input IUX 760 MI dùng cho hệ thống DCS Foxboro",
                "dvt": "Cái",
                "so_luong": 4,
                "don_gia_trinh": 13559000,
                "thanh_tien_trinh": 54236000,
                "danh_gia_ttd": "Đơn giá ERP nhập kho lần gần nhất (HĐ 115/2023) là 6.015.000 VNĐ/Cái. Đơn giá dự toán trình tăng 125% (+50.14%/năm) là chưa phù hợp. Đề nghị xem lại.",
                "phan_bien_khvt": "Đề xuất giữ nguyên. Chỉ số CPI không áp dụng cho vật tư đặc thù như vật tư đang xem xét. Đơn giá vật tư ERP thấp nhất không phải đơn giá trong vòng 12 tháng nên không thể áp dụng để lập dự toán.",
                "don_gia_thong_nhat": 13559000,
                "thanh_tien_thong_nhat": 54236000,
                "gia_tri_giam": 0,
                "co_so_thong_nhat": "Đơn giá dự toán 13.559.000 đ thấp hơn mức giá cao nhất Nhà máy từng mua năm 2024 (17.670.000 đ theo HĐ 132/2023) là -23,2%."
            },
            {
                "id": 2,
                "ma_vt": "3.34.40.292.VIE.00.000",
                "part_no": "6802",
                "ten_vt": "Mặt công tắc dùng cho 2 thiết bị",
                "dvt": "Cái",
                "so_luong": 5,
                "don_gia_trinh": 45000,
                "thanh_tien_trinh": 225000,
                "danh_gia_ttd": "Hợp đồng IMIS EVN gần nhất là 25.000 đ (Công ty Thủy điện Đồng Nai - HĐ 72/2025/HĐ-ĐN-MN ký ngày 04/09/2025). Đề nghị áp dụng theo giá IMIS.",
                "phan_bien_khvt": "Đồng ý điều chỉnh theo đơn giá HĐ IMIS của Thủy điện Đồng Nai.",
                "don_gia_thong_nhat": 25000,
                "thanh_tien_thong_nhat": 125000,
                "gia_tri_giam": 100000,
                "co_so_thong_nhat": "Thống nhất theo HĐ 72/2025/HĐ-ĐN-MN của Công ty Thủy điện Đồng Nai."
            }
        ]
    }


def save_dossier_data(data):
    os.makedirs(os.path.dirname(DATA_FILE), exist_ok=True)
    with open(DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


@app.route("/")
def index():
    dist_index = os.path.join(app.static_folder, "index.html")
    if os.path.exists(dist_index):
        return send_file(dist_index)
    return render_template("index.html")


@app.route("/api/status", methods=["GET"])
def api_status():
    info = imis_core.get_token_status_info()
    cached = imis_core.get_erp_cached_records()
    info["erp_cache_count"] = len(cached)
    return jsonify(info)


@app.route("/api/imis/config-status", methods=["GET"])
def api_imis_config_status():
    status = imis_core.get_imis_config_status()
    return jsonify(status)



@app.route("/api/refresh-token", methods=["POST"])
def api_refresh_token():
    ok, msg = imis_core.refresh_imis_token()
    info = imis_core.get_imis_config_status()
    return jsonify({"success": ok, "message": msg, "info": info})


@app.route("/api/imis/login", methods=["POST"])
def api_login_imis():
    req_data = request.get_json() or {}
    username = req_data.get("username", "").strip()
    password = req_data.get("password", "")
    remember = req_data.get("remember", True)
    if not username or not password:
        return jsonify({"success": False, "message": "Vui lòng nhập đầy đủ Tên đăng nhập và Mật khẩu"}), 400
    
    ok, msg = imis_core.login_imis(username, password, remember)
    info = imis_core.get_imis_config_status()
    return jsonify({"success": ok, "message": msg, "info": info})



@app.route("/api/dossier", methods=["GET"])
def api_get_dossier():
    return jsonify(load_dossier_data())


@app.route("/api/dossier", methods=["POST"])
def api_save_dossier():
    req_data = request.get_json()
    if not req_data:
        return jsonify({"success": False, "message": "Dữ liệu không hợp lệ"}), 400
    save_dossier_data(req_data)
    
    # Nếu có project đang active, tự động cập nhật vào file project đó
    if os.path.exists(ACTIVE_PROJECT_FILE):
        try:
            with open(ACTIVE_PROJECT_FILE, "r", encoding="utf-8") as fp:
                act = json.load(fp)
            act_id = act.get("active_id")
            if act_id:
                p_path = os.path.join(PROJECTS_DIR, act_id)
                with open(p_path, "w", encoding="utf-8") as fp:
                    json.dump(req_data, fp, ensure_ascii=False, indent=2)
        except Exception:
            pass
            
    return jsonify({"success": True, "message": "Đã lưu hồ sơ thành công"})


@app.route("/api/projects", methods=["GET"])
def api_list_projects():
    """Liệt kê danh sách tất cả các dự án đã lưu trong data/projects/."""
    os.makedirs(PROJECTS_DIR, exist_ok=True)
    res = []
    for f in os.listdir(PROJECTS_DIR):
        if f.endswith(".json"):
            fpath = os.path.join(PROJECTS_DIR, f)
            try:
                mtime = os.path.getmtime(fpath)
                mtime_str = datetime.fromtimestamp(mtime).strftime("%Y-%m-%d %H:%M:%S")
                with open(fpath, "r", encoding="utf-8") as fp:
                    pdata = json.load(fp)
                items = pdata.get("items", [])
                total_trinh = sum([it.get("thanh_tien_trinh", 0) for it in items])
                total_tn = sum([it.get("thanh_tien_thong_nhat", 0) for it in items])
                res.append({
                    "id": f,
                    "name": pdata.get("dossier_name", f.replace(".json", "")),
                    "creator": pdata.get("creator", "Nguyễn Anh Hiếu"),
                    "count": len(items),
                    "total_trinh": total_trinh,
                    "total_thong_nhat": total_tn,
                    "updated_at": mtime_str
                })
            except Exception:
                pass
    res.sort(key=lambda x: x["updated_at"], reverse=True)
    return jsonify(res)


@app.route("/api/projects/save-as", methods=["POST"])
def api_save_as_project():
    """Lưu dự án hiện tại thành một file dự án mới (Save As Project)."""
    req = request.get_json() or {}
    name = req.get("name", "").strip()
    if not name:
        return jsonify({"success": False, "message": "Tên dự án không được để trống"}), 400
        
    safe_name = re.sub(r'[\\/*?:"<>| ]', "_", name)
    filename = f"{safe_name}.json"
    fpath = os.path.join(PROJECTS_DIR, filename)
    
    data = req.get("data", {})
    data["dossier_name"] = name
    if req.get("creator"):
        data["creator"] = req["creator"]
        
    with open(fpath, "w", encoding="utf-8") as fp:
        json.dump(data, fp, ensure_ascii=False, indent=2)
        
    save_dossier_data(data)
    with open(ACTIVE_PROJECT_FILE, "w", encoding="utf-8") as fp:
        json.dump({"active_id": filename, "name": name}, fp, ensure_ascii=False, indent=2)
        
    return jsonify({
        "success": True,
        "message": f"Đã lưu thành dự án: {name}",
        "project_id": filename,
        "name": name
    })


@app.route("/api/projects/load/<filename>", methods=["GET"])
def api_load_project(filename):
    """Nạp một dự án đã lưu để tiếp tục làm việc."""
    fpath = os.path.join(PROJECTS_DIR, filename)
    if not os.path.exists(fpath):
        return jsonify({"success": False, "message": "Không tìm thấy file dự án"}), 404
        
    try:
        with open(fpath, "r", encoding="utf-8") as fp:
            data = json.load(fp)
        save_dossier_data(data)
        with open(ACTIVE_PROJECT_FILE, "w", encoding="utf-8") as fp:
            json.dump({"active_id": filename, "name": data.get("dossier_name", filename)}, fp, ensure_ascii=False, indent=2)
        return jsonify({"success": True, "dossier": data, "project_id": filename})
    except Exception as e:
        return jsonify({"success": False, "message": f"Lỗi đọc dự án: {e}"}), 500


@app.route("/api/projects/delete/<filename>", methods=["DELETE"])
def api_delete_project(filename):
    """Xóa một dự án đã lưu."""
    fpath = os.path.join(PROJECTS_DIR, filename)
    if os.path.exists(fpath):
        try:
            os.remove(fpath)
            return jsonify({"success": True, "message": "Đã xóa dự án thành công"})
        except Exception as e:
            return jsonify({"success": False, "message": f"Lỗi xóa: {e}"}), 500
    return jsonify({"success": False, "message": "Dự án không tồn tại"}), 404


def get_project_files_dir(project_id=None):
    """Lấy thư mục lưu trữ media & file đính kèm của dự án hiện tại."""
    if not project_id and os.path.exists(ACTIVE_PROJECT_FILE):
        try:
            with open(ACTIVE_PROJECT_FILE, "r", encoding="utf-8") as fp:
                act = json.load(fp)
            project_id = act.get("active_id")
        except Exception:
            pass
    if not project_id:
        project_id = "current_dossier"
    base_name = project_id.replace(".json", "")
    p_dir = os.path.join(DATA_DIR, "projects", f"{base_name}_files")
    os.makedirs(p_dir, exist_ok=True)
    return p_dir


@app.route("/api/items/<int:item_id>/upload-attachment", methods=["POST"])
def api_upload_attachment(item_id):
    """Upload file ảnh, PDF hoặc tài liệu đính kèm cho 1 mục vật tư."""
    if 'file' not in request.files:
        return jsonify({"success": False, "message": "Không có file"}), 400
    file = request.files['file']
    if not file or file.filename == '':
        return jsonify({"success": False, "message": "Chưa chọn file"}), 400
    
    p_dir = get_project_files_dir()
    item_dir = os.path.join(p_dir, f"item_{item_id}")
    os.makedirs(item_dir, exist_ok=True)
    
    safe_filename = re.sub(r'[\\/*?:"<>| ]', "_", file.filename)
    dest_path = os.path.join(item_dir, safe_filename)
    file.save(dest_path)
    
    rel_path = f"item_{item_id}/{safe_filename}"
    file_type = "image" if safe_filename.lower().endswith(('.png', '.jpg', '.jpeg', '.webp', '.gif')) else ("pdf" if safe_filename.lower().endswith('.pdf') else "other")
    
    data = load_dossier_data()
    for it in data.get("items", []):
        if it.get("id") == item_id:
            if "attachments" not in it:
                it["attachments"] = []
            it["attachments"].append({
                "name": file.filename,
                "rel_path": rel_path,
                "type": file_type,
                "size": os.path.getsize(dest_path),
                "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            })
            break
    save_dossier_data(data)
    return jsonify({"success": True, "rel_path": rel_path, "type": file_type, "name": file.filename})


@app.route("/api/items/<int:item_id>/paste-image", methods=["POST"])
def api_paste_image(item_id):
    """Lưu ảnh chụp màn hình từ Clipboard (Ctrl + V) thành file chứng cứ của mục."""
    req = request.get_json() or {}
    img_b64 = req.get("image_base64", "")
    if not img_b64:
        return jsonify({"success": False, "message": "Không có dữ liệu ảnh"}), 400
        
    if "," in img_b64:
        img_b64 = img_b64.split(",")[1]
        
    import base64
    try:
        img_bytes = base64.b64decode(img_b64)
    except Exception as e:
        return jsonify({"success": False, "message": f"Lỗi giải mã ảnh: {e}"}), 400
        
    p_dir = get_project_files_dir()
    item_dir = os.path.join(p_dir, f"item_{item_id}")
    os.makedirs(item_dir, exist_ok=True)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"clip_{timestamp}.png"
    dest_path = os.path.join(item_dir, filename)
    with open(dest_path, "wb") as f:
        f.write(img_bytes)
        
    rel_path = f"item_{item_id}/{filename}"
    data = load_dossier_data()
    for it in data.get("items", []):
        if it.get("id") == item_id:
            if "attachments" not in it:
                it["attachments"] = []
            it["attachments"].append({
                "name": filename,
                "rel_path": rel_path,
                "type": "image",
                "size": len(img_bytes),
                "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            })
            break
    save_dossier_data(data)
    return jsonify({"success": True, "rel_path": rel_path, "type": "image", "name": filename})


@app.route("/api/items/<int:item_id>/add-link", methods=["POST"])
def api_add_link(item_id):
    """Lưu link URL internet (trang mua sắm công, website chính hãng) cho mục."""
    req = request.get_json() or {}
    url = req.get("url", "").strip()
    title = req.get("title", "").strip() or url
    if not url:
        return jsonify({"success": False, "message": "Chưa nhập URL"}), 400
        
    data = load_dossier_data()
    for it in data.get("items", []):
        if it.get("id") == item_id:
            if "links" not in it:
                it["links"] = []
            it["links"].append({
                "url": url,
                "title": title,
                "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            })
            break
    save_dossier_data(data)
    return jsonify({"success": True})


@app.route("/api/project-files/<path:rel_path>", methods=["GET"])
def api_serve_project_file(rel_path):
    """Phục vụ file media/PDF của dự án để hiển thị trực tiếp trên trình duyệt."""
    p_dir = get_project_files_dir()
    full_path = os.path.join(p_dir, rel_path)
    if os.path.exists(full_path):
        return send_file(full_path)
    return jsonify({"error": "File not found"}), 404


@app.route("/api/items/<int:item_id>/ai-markdown", methods=["POST"])
def api_save_ai_markdown(item_id):
    """Lưu trữ bài phân tích AI dạng Markdown cho 1 mục vật tư."""
    req = request.get_json() or {}
    md_content = req.get("markdown", "")
    
    data = load_dossier_data()
    for it in data.get("items", []):
        if it.get("id") == item_id:
            it["ai_analysis_md"] = md_content
            break
    save_dossier_data(data)
    
    p_dir = get_project_files_dir()
    item_dir = os.path.join(p_dir, f"item_{item_id}")
    os.makedirs(item_dir, exist_ok=True)
    md_file = os.path.join(item_dir, "phan_tich_ai.md")
    with open(md_file, "w", encoding="utf-8") as f:
        f.write(md_content)
        
    return jsonify({"success": True, "message": "Đã lưu bản phân tích AI"})


@app.route("/api/items/<int:item_id>/delete-attachment", methods=["POST"])
def api_delete_attachment(item_id):
    """Xóa 1 file đính kèm khỏi mục vật tư."""
    req = request.get_json() or {}
    rel_path = req.get("rel_path", "")
    
    data = load_dossier_data()
    for it in data.get("items", []):
        if it.get("id") == item_id:
            it["attachments"] = [a for a in it.get("attachments", []) if a.get("rel_path") != rel_path]
            break
    save_dossier_data(data)
    
    p_dir = get_project_files_dir()
    full_path = os.path.join(p_dir, rel_path)
    if os.path.exists(full_path):
        try:
            os.remove(full_path)
        except Exception:
            pass
    return jsonify({"success": True})


@app.route("/api/quotes/scan", methods=["POST"])
def api_scan_quotes():
    """Quét thư mục chứa các file PDF báo giá của các nhà thầu."""
    req = request.get_json() or {}
    folder = req.get("folder_path") or quote_matcher.DEFAULT_QUOTES_DIR
    force_rescan = bool(req.get("force_rescan", False))
    res = quote_matcher.scan_quotation_folder(folder, force_rescan=force_rescan)
    return jsonify(res)


@app.route("/api/quotes/view-pdf")
def api_view_quote_pdf():
    """Xem trực tiếp file PDF báo giá gốc trên trình duyệt để kiểm tra tổng thành tiền."""
    fpath = request.args.get("path") or ""
    filename = request.args.get("filename") or ""

    # 1. Kiểm tra fpath chuẩn hóa
    if fpath:
        # Chuẩn hóa đường dẫn
        norm_path = os.path.normpath(fpath.strip())
        if os.path.exists(norm_path) and os.path.isfile(norm_path):
            return send_file(norm_path, mimetype="application/pdf")
            
    # 2. Nếu path bị lỗi hoặc mất gạch chéo do JS escape, tìm theo tên file
    target_fn = filename.strip() if filename else ""
    if not target_fn and fpath:
        # Nếu fpath bị dính liền nhưng có đuôi .pdf, trích xuất tên file
        m = re.search(r'([^\\/]+\.pdf)$', fpath, re.IGNORECASE)
        if m:
            target_fn = m.group(1)

    if target_fn:
        quotes_dir = quote_matcher.DEFAULT_QUOTES_DIR
        candidate = os.path.join(quotes_dir, target_fn)
        if os.path.exists(candidate) and os.path.isfile(candidate):
            return send_file(candidate, mimetype="application/pdf")

        # Quét thư mục tìm file khớp tên không phân biệt hoa thường
        for root, dirs, files in os.walk(quotes_dir):
            for f in files:
                if f.lower() == target_fn.lower() or target_fn.lower().endswith(f.lower()) or f.lower().endswith(target_fn.lower()):
                    full_match = os.path.join(root, f)
                    if os.path.exists(full_match) and os.path.isfile(full_match):
                        return send_file(full_match, mimetype="application/pdf")

    return f"Không tìm thấy file PDF: {filename or fpath}", 404


def get_project_quote_overrides():
    """Lấy dữ liệu hiệu chỉnh báo giá thủ công đã lưu của dự án."""
    p_dir = get_project_files_dir()
    fpath = os.path.join(p_dir, "quote_overrides.json")
    if os.path.exists(fpath):
        try:
            with open(fpath, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {}
    return {}


@app.route("/api/quotes/save-edited-quote", methods=["POST"])
@app.route("/api/quotes/save-override", methods=["POST"])
def api_save_edited_quote():
    """Lưu dữ liệu báo giá đã được người dùng chỉnh sửa/bổ sung và tính lại thành tiền."""
    req = request.get_json() or {}
    quote_obj = req.get("quote", {})
    filename = req.get("filename") or quote_obj.get("filename")
    items = req.get("items") or quote_obj.get("items") or []
    total_amount = float(req.get("total_amount") or quote_obj.get("total_amount") or 0)
    
    if not filename:
        return jsonify({"success": False, "message": "Thiếu tên file báo giá"}), 400

    p_dir = get_project_files_dir()
    os.makedirs(p_dir, exist_ok=True)
    fpath = os.path.join(p_dir, "quote_overrides.json")
    overrides = get_project_quote_overrides()
    overrides[filename] = {
        "items": items,
        "item_count": len(items),
        "total_amount": total_amount,
        "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    }
    with open(fpath, "w", encoding="utf-8") as f:
        json.dump(overrides, f, ensure_ascii=False, indent=2)

    # Xóa cache để lượt quét tới áp dụng override mới ngay lập tức
    quote_matcher.clear_quotes_cache()

    return jsonify({"success": True, "message": f"Đã lưu thành công {len(items)} dòng dữ liệu hiệu chỉnh cho báo giá [{filename}]!"})


@app.route("/api/quotes/get-full-quote", methods=["GET", "POST"])
@app.route("/api/quotes/item-data", methods=["GET", "POST"])
def api_get_full_quote():
    """Lấy dữ liệu đầy đủ tất cả các dòng đã pandas hóa của 1 file báo giá."""
    req = request.get_json(silent=True) or {}
    filename = req.get("filename") or request.args.get("filename")
    folder = req.get("folder_path") or request.args.get("folder_path")
    
    if not folder:
        p_dir = get_project_files_dir()
        approved_file = os.path.join(p_dir, "bao_gia_project.json")
        if os.path.exists(approved_file):
            try:
                with open(approved_file, "r", encoding="utf-8") as f:
                    folder = json.load(f).get("folder_nguon")
            except Exception:
                pass
                
    folder = folder or quote_matcher.DEFAULT_QUOTES_DIR
    overrides = get_project_quote_overrides()
    scanned = quote_matcher.scan_quotation_folder(folder, overrides=overrides)
    for q in scanned.get("quotes", []):
        if q["filename"] == filename:
            return jsonify({"success": True, "quote": q, "data": q})
    return jsonify({"success": False, "message": f"Không tìm thấy báo giá: {filename}"}), 404


@app.route("/api/quotes/dossier", methods=["GET", "POST"])
def api_quotes_dossier():
    """Lấy danh sách tất cả các file trong thư mục báo giá, phân loại và trạng thái phê duyệt."""
    req = request.get_json() if request.method == "POST" else {}
    folder = (req or {}).get("folder_path") or request.args.get("folder_path") or quote_matcher.DEFAULT_QUOTES_DIR
    overrides = get_project_quote_overrides()
    res = quote_matcher.scan_quotation_folder(folder, overrides=overrides)
    
    p_dir = get_project_files_dir()
    approved_file = os.path.join(p_dir, "bao_gia_project.json")
    approved_data = None
    if os.path.exists(approved_file):
        try:
            with open(approved_file, "r", encoding="utf-8") as f:
                approved_data = json.load(f)
        except Exception:
            pass
            
    res["approved_data"] = approved_data
    res["is_approved"] = approved_data is not None
    return jsonify(res)


@app.route("/api/quotes/approve-all", methods=["POST"])
def api_quotes_approve_all():
    """Phê duyệt bộ dữ liệu báo giá đã số hóa vào CSDL chính thức của dự án."""
    req = request.get_json() or {}
    folder = req.get("folder_path") or quote_matcher.DEFAULT_QUOTES_DIR
    overrides = get_project_quote_overrides()
    res = quote_matcher.scan_quotation_folder(folder, overrides=overrides)
    
    p_dir = get_project_files_dir()
    os.makedirs(p_dir, exist_ok=True)
    approved_file = os.path.join(p_dir, "bao_gia_project.json")
    
    save_payload = {
        "thoi_gian_duyet": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "nguoi_duyet": "Nguyễn Anh Hiếu - Tổ Thẩm định",
        "folder_nguon": folder,
        "tong_so_nha_thau": len(res.get("quotes", [])),
        "tong_so_muc": sum(q.get("item_count", 0) for q in res.get("quotes", [])),
        "tong_gia_tri": sum(q.get("total_amount", 0) for q in res.get("quotes", [])),
        "danh_sach_bao_gia": res.get("quotes", []),
        "scans": res.get("scans", []),
        "docs": res.get("docs", [])
    }
    
    with open(approved_file, "w", encoding="utf-8") as f:
        json.dump(save_payload, f, ensure_ascii=False, indent=2)
        
    return jsonify({"success": True, "message": "Đã phê duyệt toàn bộ dữ liệu báo giá vào CSDL Dự án thành công!", "approved_summary": save_payload})


@app.route("/api/quotes/match-item", methods=["POST"])
def api_match_item_quote():
    """Đối chiếu đơn giá trình của 1 mục với các báo giá gốc trong thư mục."""
    req = request.get_json() or {}
    item = req.get("item", {})
    folder = req.get("folder_path")
    
    if not folder:
        p_dir = get_project_files_dir()
        approved_file = os.path.join(p_dir, "bao_gia_project.json")
        if os.path.exists(approved_file):
            try:
                with open(approved_file, "r", encoding="utf-8") as f:
                    folder = json.load(f).get("folder_nguon")
            except Exception:
                pass
                
    folder = folder or quote_matcher.DEFAULT_QUOTES_DIR
    force_rescan = bool(req.get("force_rescan", False))
    
    overrides = get_project_quote_overrides()
    scanned = quote_matcher.scan_quotation_folder(folder, overrides=overrides, force_rescan=force_rescan)
    match_result = quote_matcher.match_item_in_quotes(item, scanned)
    return jsonify(match_result)


@app.route("/api/quotes/match-all-dossier-items", methods=["GET", "POST"])
def api_match_all_dossier_items():
    """Tự động đối chiếu toàn bộ danh mục vật tư trong dự án với các file Báo giá gốc."""
    dossier = load_dossier_data()
    items = dossier.get("items", [])
    
    p_dir = get_project_files_dir()
    approved_file = os.path.join(p_dir, "bao_gia_project.json")
    folder = None
    if os.path.exists(approved_file):
        try:
            with open(approved_file, "r", encoding="utf-8") as f:
                folder = json.load(f).get("folder_nguon")
        except Exception:
            pass
    folder = folder or quote_matcher.DEFAULT_QUOTES_DIR
    
    overrides = get_project_quote_overrides()
    scanned = quote_matcher.scan_quotation_folder(folder, overrides=overrides)
    
    results = {}
    for idx, item in enumerate(items):
        item_id = item.get("id", idx + 1)
        res = quote_matcher.match_item_in_quotes(item, scanned)
        min_vendor = ""
        if res.get("matches"):
            min_vendor = res["matches"][0].get("company", "")
        results[item_id] = {
            "lowest_price": res.get("min_price"),
            "lowest_vendor": min_vendor,
            "co_so_don_gia": res.get("summary_text") or item.get("co_so_thong_nhat") or item.get("danh_gia_ttd") or item.get("ghi_chu") or "",
            "matches_count": len(res.get("matches", []))
        }
        
    return jsonify({"success": True, "results": results})


@app.route("/api/erp/config-status", methods=["GET"])
def api_erp_config_status():
    """Kiểm tra trạng thái CSDL ERP khi ứng dụng khởi chạy."""
    status = imis_core.get_erp_config_status()
    return jsonify(status)


@app.route("/api/erp/preview-columns", methods=["POST"])
def api_erp_preview_columns():
    """Đọc tiêu đề cột của file Excel ERP để hiển thị trên UI chọn ánh xạ."""
    req = request.get_json() or {}
    file_path = req.get("file_path", "").strip()
    res = imis_core.get_excel_headers(file_path)
    return jsonify(res)


@app.route("/api/erp/upload", methods=["POST"])
def api_erp_upload_file():
    """Tải lên file Excel CSDL ERP mới từ giao diện web."""
    if 'file' not in request.files:
        return jsonify({"success": False, "message": "Không tìm thấy file"}), 400
    file = request.files['file']
    if not file.filename or not file.filename.lower().endswith(('.xlsx', '.xls')):
        return jsonify({"success": False, "message": "Chỉ chấp nhận file Excel (.xlsx, .xls)"}), 400
        
    config_dir = os.path.join(DATA_DIR, "config")
    os.makedirs(config_dir, exist_ok=True)
    dest_path = os.path.join(config_dir, "ERP_uploaded.xlsx")
    file.save(dest_path)
    
    headers_res = imis_core.get_excel_headers(dest_path)
    headers_res["uploaded_path"] = dest_path
    return jsonify(headers_res)


@app.route("/api/erp/save-config", methods=["POST"])
def api_erp_save_config():
    """Lưu cấu hình vị trí file Excel ERP và mapping 13 cột pháp lý."""
    req = request.get_json() or {}
    file_path = req.get("file_path", "").strip()
    mapping = req.get("mapping", {})
    header_row = int(req.get("header_row", 1))
    
    if not file_path or not os.path.exists(file_path):
        return jsonify({"success": False, "message": f"Không tìm thấy file Excel: {file_path}"}), 400
        
    records = imis_core.save_erp_mapping_config(file_path, mapping, header_row=header_row)
    return jsonify({
        "success": True,
        "message": f"Đã nạp thành công CSDL ERP với {len(records)} bản ghi hợp đồng!",
        "count": len(records)
    })


@app.route("/api/erp/search", methods=["POST"])
def api_erp_search():
    """Tra cứu lịch sử mua sắm CSDL Kế toán ERP Vĩnh Tân 4."""
    req = request.get_json() or {}
    keyword = req.get("keyword", "").strip()
    ma_vt = req.get("ma_vt", "").strip()
    item = req.get("item") or {}
    dg_trinh = float(req.get("dg_trinh") or item.get("don_gia_trinh") or 0)
    selected_record = req.get("selected_record")
    is_manual = req.get("is_manual", False)
    
    min_score = 0 if is_manual else 60
    
    results = imis_core.search_erp_baseline(keyword, ma_vt=ma_vt, min_score=min_score)
    if not results and not is_manual and keyword:
        clean_kw = keyword.split("\n")[0].split("-")[0].strip()
        results = imis_core.search_erp_baseline(clean_kw, min_score=40)
        
    cfg = imis_core.load_erp_mapping_config()
    mapping = cfg.get("mapping", {})
    
    use_average = req.get("use_average", False) or selected_record == "AVERAGE"
    summary_data = imis_core.generate_erp_summary_text(item, results, dg_trinh=dg_trinh, selected_record=selected_record, use_average=use_average)
    return jsonify({
        "success": True,
        "results": results,
        "mapping": mapping,
        "summary": summary_data,
        "summary_text": summary_data.get("summary_text", "")
    })


@app.route("/api/quotes/browse-folders", methods=["GET", "POST"])
def api_quotes_browse_folders():
    """Duyệt danh sách thư mục con để hiển thị cây thư mục trên UI."""
    req = request.get_json(silent=True) or {}
    base_path = req.get("path", "").strip() or request.args.get("path", "").strip() or "D:\\"
    if not os.path.exists(base_path):
        base_path = os.path.dirname(base_path) if os.path.dirname(base_path) else "C:\\"
    
    subdirs = []
    try:
        if os.path.isdir(base_path):
            for entry in os.listdir(base_path):
                full_p = os.path.join(base_path, entry)
                if os.path.isdir(full_p) and not entry.startswith('.') and not entry.startswith('$'):
                    subdirs.append({"name": entry, "path": full_p})
    except Exception:
        pass
    
    parent_p = os.path.dirname(base_path) if os.path.dirname(base_path) != base_path else None
    return jsonify({
        "success": True,
        "current_path": base_path,
        "parent_path": parent_p,
        "subdirs": subdirs[:40]
    })


@app.route("/api/quotes/native-browse-folder", methods=["GET", "POST"])
def api_quotes_native_browse_folder():
    """Mở cửa sổ Windows Explorer Native Folder Picker Dialog chuẩn của hệ điều hành."""
    try:
        import tkinter as tk
        from tkinter import filedialog
        root = tk.Tk()
        root.withdraw()
        root.attributes('-topmost', True)
        selected_dir = filedialog.askdirectory(title="Chọn thư mục chứa các file Báo Giá Gốc (PDF)")
        root.destroy()
        if selected_dir:
            norm_path = os.path.normpath(selected_dir)
            return jsonify({"success": True, "folder_path": norm_path})
        return jsonify({"success": False, "message": "Người dùng đã hủy chọn thư mục"})
    except Exception as e:
        return jsonify({"success": False, "message": f"Lỗi mở cửa sổ Windows: {str(e)}"})


@app.route("/api/quotes/attach-matched-pdf", methods=["POST"])
def api_attach_matched_pdf():
    """Tự động đính kèm file PDF báo giá gốc vào Kho chứng cứ của mục."""
    req = request.get_json() or {}
    item_id = req.get("item_id")
    pdf_path = req.get("pdf_path")
    if not item_id or not pdf_path or not os.path.exists(pdf_path):
        return jsonify({"success": False, "message": "File báo giá không tồn tại"}), 400
        
    filename = os.path.basename(pdf_path)
    p_dir = get_project_files_dir()
    item_dir = os.path.join(p_dir, f"item_{item_id}")
    os.makedirs(item_dir, exist_ok=True)
    
    dest_path = os.path.join(item_dir, filename)
    import shutil
    shutil.copy2(pdf_path, dest_path)
    
    rel_path = f"item_{item_id}/{filename}"
    data = load_dossier_data()
    for it in data.get("items", []):
        if it.get("id") == item_id:
            if "attachments" not in it:
                it["attachments"] = []
            if not any(a.get("rel_path") == rel_path for a in it["attachments"]):
                it["attachments"].append({
                    "name": filename,
                    "rel_path": rel_path,
                    "type": "pdf",
                    "size": os.path.getsize(dest_path),
                    "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                })
            break
    save_dossier_data(data)
    return jsonify({"success": True, "rel_path": rel_path, "filename": filename})


@app.route("/api/msc/update-curl", methods=["POST"])
def api_msc_update_curl():
    """Cập nhật và kích hoạt phiên Mua Sắm Công bằng chuỗi cURL dán từ Chrome DevTools."""
    req = request.get_json() or {}
    curl_str = req.get("curl_command", "").strip()
    sess, err = msc_matcher.parse_curl_command(curl_str)
    if err:
        return jsonify({"success": False, "message": err}), 400
        
    test_res = msc_matcher.test_msc_connection(sess)
    if not test_res["active"]:
        return jsonify({"success": False, "message": test_res["message"]}), 400
        
    msc_matcher.save_msc_session(sess)
    return jsonify({"success": True, "message": "Đã cập nhật và kích hoạt phiên Mua Sắm Công thành công!"})


@app.route("/api/msc/status", methods=["GET"])
def api_msc_status():
    """Kiểm tra trạng thái kết nối Mua Sắm Công."""
    return jsonify(msc_matcher.test_msc_connection())


@app.route("/api/msc/search", methods=["POST"])
@app.route("/api/msc/search-item", methods=["POST"])
def api_msc_search():
    """Tra cứu đơn giá trúng thầu Mua Sắm Công cho 1 mục và tự động lưu vết chứng cứ."""
    req = request.get_json() or {}
    keyword = req.get("keyword", "").strip()
    item = req.get("item", {})
    save_evidence = req.get("save_evidence", True)
    page_num = int(req.get("page_number") or req.get("page_num") or 0)
    page_sz = int(req.get("page_size") or req.get("page_sz") or 20)
    
    res = msc_matcher.search_muasamcong(keyword, page_number=page_num, page_size=page_sz)
    if not res.get("success"):
        return jsonify(res)
        
    comp = msc_matcher.analyze_msc_comparison(item, res)
    
    # Tự động lưu chứng cứ vào file chung_cu_muasamcong.json
    item_id = item.get("id")
    if save_evidence and item_id:
        p_dir = get_project_files_dir()
        item_dir = os.path.join(p_dir, f"item_{item_id}")
        os.makedirs(item_dir, exist_ok=True)
        
        evidence_data = {
            "item_id": item_id,
            "tu_khoa_tra_cuu": keyword,
            "thoi_gian_tra_cuu": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "nguon": "Mạng Đấu thầu Quốc gia (muasamcong.mpi.gov.vn)",
            "don_gia_trinh": item.get("don_gia_trinh", 0),
            "don_gia_tham_chieu": comp.get("min_price", 0),
            "chenh_lech_so_tien": comp.get("diff_amt", 0),
            "chenh_lech_phan_tram": comp.get("diff_pct", 0),
            "danh_sach_ket_qua": comp.get("items", [])
        }
        with open(os.path.join(item_dir, "chung_cu_muasamcong.json"), "w", encoding="utf-8") as f:
            json.dump(evidence_data, f, ensure_ascii=False, indent=2)
            
    return jsonify({"success": True, "analysis": comp})


@app.route("/api/evidence/save-step", methods=["POST"])
def api_save_evidence_step():
    """Lưu bằng chứng tiến trình tra cứu cho ERP, IMIS hoặc Báo giá."""
    req = request.get_json() or {}
    item_id = req.get("item_id")
    step_type = req.get("step_type")
    payload = req.get("payload", {})
    if not item_id or not step_type:
        return jsonify({"success": False, "message": "Thiếu thông tin"}), 400
        
    p_dir = get_project_files_dir()
    item_dir = os.path.join(p_dir, f"item_{item_id}")
    os.makedirs(item_dir, exist_ok=True)
    
    fname = f"chung_cu_{step_type}.json"
    payload["thoi_gian_luu"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with open(os.path.join(item_dir, fname), "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
        
    return jsonify({"success": True, "filename": fname})


@app.route("/api/evidence/status/<int:item_id>", methods=["GET"])
def api_evidence_status(item_id):
    """Kiểm tra xem mục này đã có các chứng cứ nào được lưu."""
    p_dir = get_project_files_dir()
    item_dir = os.path.join(p_dir, f"item_{item_id}")
    k1 = os.path.exists(os.path.join(item_dir, "chung_cu_quotes.json"))
    k2 = os.path.exists(os.path.join(item_dir, "chung_cu_erp.json"))
    k3 = os.path.exists(os.path.join(item_dir, "chung_cu_imis.json"))
    k4 = os.path.exists(os.path.join(item_dir, "chung_cu_muasamcong.json"))
    return jsonify({
        "has_quotes": k1,
        "has_erp": k2,
        "has_imis": k3,
        "has_msc": k4,
        "done_count": sum([1 if x else 0 for x in [k1, k2, k3, k4]])
    })


@app.route("/api/evidence/all-status", methods=["GET"])
def api_evidence_all_status():
    """Kiểm tra tiến độ 4 khối của toàn bộ các mục trong dự án."""
    p_dir = get_project_files_dir()
    status_map = {}
    if os.path.exists(p_dir):
        for entry in os.listdir(p_dir):
            if entry.startswith("item_"):
                try:
                    i_id = int(entry.replace("item_", ""))
                    i_dir = os.path.join(p_dir, entry)
                    if os.path.isdir(i_dir):
                        k1 = os.path.exists(os.path.join(i_dir, "chung_cu_quotes.json"))
                        k2 = os.path.exists(os.path.join(i_dir, "chung_cu_erp.json"))
                        k3 = os.path.exists(os.path.join(i_dir, "chung_cu_imis.json"))
                        k4 = os.path.exists(os.path.join(i_dir, "chung_cu_muasamcong.json"))
                        status_map[str(i_id)] = {
                            "has_quotes": k1,
                            "has_erp": k2,
                            "has_imis": k3,
                            "has_msc": k4,
                            "done_count": sum([1 if x else 0 for x in [k1, k2, k3, k4]])
                        }
                except Exception:
                    pass
    return jsonify(status_map)


@app.route("/api/search-item-sources", methods=["POST"])
def api_search_sources():
    req = request.get_json() or {}
    kw = req.get("keyword", "").strip()
    tu_ngay = req.get("tu_ngay", "2023-01-01")
    den_ngay = req.get("den_ngay")
    ma_vt = req.get("ma_vt", "")
    item = req.get("item", kw)
    dg_trinh = float(req.get("dg_trinh") or 0)
    selected_record = req.get("selected_record")
    use_average = req.get("use_average", False)

    if not kw:
        return jsonify({"imis": [], "erp": [], "summary": None, "summary_text": ""})

    result = imis_core.search_item_sources(kw, tu_ngay=tu_ngay, den_ngay=den_ngay, ma_vt=ma_vt)
    imis_recs = result.get("imis", [])
    
    used_kw = result.get("used_keyword") or kw
    summary_data = imis_core.generate_imis_summary_text(
        item, imis_recs, dg_trinh=dg_trinh, selected_record=selected_record, use_average=use_average, tu_ngay=tu_ngay, den_ngay=den_ngay, search_keyword=used_kw
    )
    result["summary"] = summary_data
    result["summary_text"] = summary_data.get("summary_text", "")
    return jsonify(result)


@app.route("/api/import-excel", methods=["POST"])
def api_import_excel():
    if 'file' not in request.files:
        return jsonify({"success": False, "message": "Không tìm thấy file"}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({"success": False, "message": "Chưa chọn file"}), 400
        
    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active
        
        items = []
        item_id = 1
        rows_list = list(ws.iter_rows(values_only=True))
        
        # Tự động tìm dòng Tiêu đề (Header)
        start_row_idx = 0
        is_13_cols_format = False
        for idx, r in enumerate(rows_list):
            if not r:
                continue
            r_str = " ".join([str(c or "").upper() for c in r])
            if "THÔNG SỐ KỸ THUẬT" in r_str or "ĐƠN GIÁ MIN" in r_str or ("PYCVT" in r_str and "MÃ ERP" in r_str):
                start_row_idx = idx + 1
                is_13_cols_format = True
                break
            elif ("TÊN" in r_str and ("QUY CÁCH" in r_str or "VẬT TƯ" in r_str)) or ("MÃ VẬT TƯ" in r_str and "STT" in r_str):
                start_row_idx = idx + 1
                break
                
        for row in rows_list[start_row_idx:]:
            if not row or not any(row):
                continue

            if is_13_cols_format:
                pycvt = str(row[1] if len(row) > 1 else "").strip()
                ten_vt_goc = str(row[2] if len(row) > 2 else "").strip()
                thong_so_kt = str(row[3] if len(row) > 3 else "").strip()
                dvt = str(row[4] if len(row) > 4 else "Cái").strip()
                try:
                    sl = float(row[5] if len(row) > 5 else 1)
                except:
                    sl = 1
                hsx_xx = str(row[6] if len(row) > 6 else "").strip()
                ma_erp = str(row[7] if len(row) > 7 else "").strip()
                try:
                    dg_trinh = float(row[8] if len(row) > 8 else 0)
                except:
                    dg_trinh = 0
                try:
                    tt_trinh = float(row[9] if len(row) > 9 else round(sl * dg_trinh, 0))
                except:
                    tt_trinh = round(sl * dg_trinh, 0)
                thue = str(row[10] if len(row) > 10 else "").strip()
                try:
                    tien_thue = float(row[11] if len(row) > 11 else 0)
                except:
                    tien_thue = 0
                ghi_chu = str(row[12] if len(row) > 12 else "").strip()

                if not ten_vt_goc and not thong_so_kt:
                    continue
                if ten_vt_goc.upper() in ("TÊN VẬT TƯ", "STT"):
                    continue

                full_name = f"{ten_vt_goc} - {thong_so_kt}" if (ten_vt_goc and thong_so_kt) else (ten_vt_goc or thong_so_kt)
                part_no = thong_so_kt or ma_erp

                items.append({
                    "id": item_id,
                    "pycvt": pycvt,
                    "ma_vt": ma_erp,
                    "part_no": part_no,
                    "ten_vt": full_name,
                    "ten_vt_goc": ten_vt_goc,
                    "thong_so_kt": thong_so_kt,
                    "hsx_xx": hsx_xx,
                    "dvt": dvt,
                    "so_luong": sl,
                    "don_gia_trinh": dg_trinh,
                    "thanh_tien_trinh": tt_trinh,
                    "thue": thue,
                    "tien_thue": tien_thue,
                    "danh_gia_ttd": "",
                    "phan_bien_khvt": "",
                    "don_gia_thong_nhat": dg_trinh,
                    "thanh_tien_thong_nhat": tt_trinh,
                    "gia_tri_giam": 0,
                    "co_so_thong_nhat": "",
                    "ghi_chu": ghi_chu
                })
            else:
                ten_vt = str(row[2] if len(row) > 2 else row[1] or "").strip()
                if not ten_vt or ten_vt.upper() in ("TÊN QUY CÁCH", "TÊN VẬT TƯ", "TÊN QUY CÁCH KỸ THUẬT VẬT TƯ"):
                    continue
                    
                ma_vt = str(row[1] if len(row) > 1 else "").strip()
                dvt = str(row[3] if len(row) > 3 else "Cái").strip()
                try:
                    sl = float(row[4] if len(row) > 4 else 1)
                except:
                    sl = 1
                try:
                    dg_trinh = float(row[5] if len(row) > 5 else 0)
                except:
                    dg_trinh = 0
                    
                tt_trinh = round(sl * dg_trinh, 0)
                
                # Đọc các cột ý kiến nếu có
                dg_ttd = str(row[7] if len(row) > 7 else "").strip()
                pb_khvt = str(row[8] if len(row) > 8 else "").strip()
                try:
                    dg_tn = float(row[9] if len(row) > 9 else dg_trinh)
                except:
                    dg_tn = dg_trinh
                tt_tn = round(sl * dg_tn, 0)
                gia_giam = max(0, tt_trinh - tt_tn)
                co_so_tn = str(row[12] if len(row) > 12 else "").strip()
                
                items.append({
                    "id": item_id,
                    "pycvt": "",
                    "ma_vt": ma_vt,
                    "part_no": ma_vt,
                    "ten_vt": ten_vt,
                    "ten_vt_goc": ten_vt,
                    "thong_so_kt": "",
                    "hsx_xx": "",
                    "dvt": dvt,
                    "so_luong": sl,
                    "don_gia_trinh": dg_trinh,
                    "thanh_tien_trinh": tt_trinh,
                    "danh_gia_ttd": dg_ttd,
                    "phan_bien_khvt": pb_khvt,
                    "don_gia_thong_nhat": dg_tn,
                    "thanh_tien_thong_nhat": tt_tn,
                    "gia_tri_giam": gia_giam,
                    "co_so_thong_nhat": co_so_tn
                })
            item_id += 1
            
        data = load_dossier_data()
        data["items"] = items
        data["dossier_name"] = os.path.splitext(file.filename)[0]
        save_dossier_data(data)
        return jsonify({"success": True, "dossier": data, "count": len(items)})
    except Exception as e:
        return jsonify({"success": False, "message": f"Lỗi đọc file Excel: {e}"}), 500


@app.route("/api/export-excel", methods=["GET"])
def api_export_excel():
    data = load_dossier_data()
    items = data.get("items", [])
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tham_Dinh_Du_Toan"
    
    # Header Styles
    font_title = Font(name="Times New Roman", size=13, bold=True, color="003366")
    font_header = Font(name="Times New Roman", size=10, bold=True, color="FFFFFF")
    fill_header_ttd = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    fill_header_khvt = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
    fill_header_res = PatternFill(start_color="E65100", end_color="E65100", fill_type="solid")
    border_thin = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )
    
    # Title Rows
    ws.append([data.get("dossier_name", "BẢNG TỔNG HỢP Ý KIẾN THẨM ĐỊNH DỰ TOÁN")])
    ws.append(["Tổ Thẩm định Dự toán - Nhà máy Nhiệt điện Vĩnh Tân 4"])
    ws.append([])
    
    headers = [
        "STT", "Mã Vật Tư", "Tên Quy Cách Kỹ Thuật", "ĐVT", "Số Lượng",
        "Đơn Giá Đề Nghị (Trình)", "Thành Tiền Đề Nghị",
        "ĐÁNH GIÁ CỦA TỔ THẨM ĐỊNH (TTĐ)",
        "Ý KIẾN PHẢN BIỆN CỦA PHÒNG KHVT",
        "Đơn Giá Thống Nhất", "Thành Tiền Thống Nhất", "Giá Trị Giảm",
        "Cơ Sở Thống Nhất"
    ]
    ws.append(headers)
    
    row_header_idx = 4
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=row_header_idx, column=col_idx)
        cell.font = font_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if col_idx in (8,):
            cell.fill = fill_header_ttd
        elif col_idx in (9,):
            cell.fill = fill_header_khvt
        elif col_idx in (10, 11, 12, 13):
            cell.fill = fill_header_res
        else:
            cell.fill = fill_header_ttd

    font_data = Font(name="Times New Roman", size=10)
    for idx, it in enumerate(items, 1):
        dg_trinh = it.get("don_gia_trinh", 0)
        sl = it.get("so_luong", 0)
        tt_trinh = round(sl * dg_trinh, 0)
        dg_tn = it.get("don_gia_thong_nhat", dg_trinh)
        tt_tn = round(sl * dg_tn, 0)
        giam = max(0, tt_trinh - tt_tn)
        
        row_vals = [
            idx,
            it.get("ma_vt", ""),
            it.get("ten_vt", ""),
            it.get("dvt", ""),
            sl,
            dg_trinh,
            tt_trinh,
            it.get("danh_gia_ttd", ""),
            it.get("phan_bien_khvt", ""),
            dg_tn,
            tt_tn,
            giam,
            it.get("co_so_thong_nhat", "")
        ]
        ws.append(row_vals)
        cur_row = row_header_idx + idx
        for col_idx in range(1, len(headers) + 1):
            c = ws.cell(row=cur_row, column=col_idx)
            c.font = font_data
            c.border = border_thin
            if col_idx in (5, 6, 7, 10, 11, 12):
                c.number_format = '#,##0'
            if col_idx in (1, 4):
                c.alignment = Alignment(horizontal="center", vertical="top")
            elif col_idx in (8, 9, 13):
                c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            else:
                c.alignment = Alignment(vertical="top")

    # Tự động chỉnh độ rộng cột
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 36
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 8
    ws.column_dimensions['F'].width = 16
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 38
    ws.column_dimensions['I'].width = 38
    ws.column_dimensions['J'].width = 16
    ws.column_dimensions['K'].width = 18
    ws.column_dimensions['L'].width = 16
    ws.column_dimensions['M'].width = 38
    
    export_path = os.path.join(os.path.abspath(os.path.dirname(__file__)), "data", "Bang_Tham_Dinh_Du_Toan.xlsx")
    wb.save(export_path)
    return send_file(export_path, as_attachment=True, download_name="Bang_Tham_Dinh_Du_Toan.xlsx")


@app.route("/api/download-template", methods=["GET"])
def api_download_template():
    """Xuất file Excel mẫu chuẩn 13 cột theo đúng quy định mua sắm của EVN Vĩnh Tân 4."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mau_Du_Toan"

    # Font & Fills
    font_title = Font(name="Times New Roman", size=13, bold=True, color="003366")
    font_sub = Font(name="Times New Roman", size=10, italic=True, color="555555")
    font_header = Font(name="Times New Roman", size=10, bold=True, color="FFFFFF")
    fill_header = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    border_thin = Border(
        left=Side(style='thin', color='B0B0B0'),
        right=Side(style='thin', color='B0B0B0'),
        top=Side(style='thin', color='B0B0B0'),
        bottom=Side(style='thin', color='B0B0B0')
    )

    # Title Banner
    ws.append(["MẪU BẢNG DỰ TOÁN ĐỀ NGHỊ MUA SẮM & THẨM ĐỊNH GIÁ VẬT TƯ"])
    ws.cell(row=1, column=1).font = font_title
    ws.append(["(Điền danh mục theo đúng các cột dưới đây, sau đó dùng nút 'Nạp Excel Dự Toán' trên ứng dụng để nhập dữ liệu)"])
    ws.cell(row=2, column=1).font = font_sub
    ws.append([])

    # 13 CỘT CHUẨN THEO ĐÚNG YÊU CẦU:
    headers = [
        "STT", "PYCVT", "Tên vật tư", "Thông số kỹ thuật", "ĐVT",
        "Số lượng mua sắm", "HSX/XX", "Mã ERP", "Đơn giá min",
        "Thành tiền", "Thuế", "Tiền thuế", "Ghi chú"
    ]
    ws.append(headers)
    row_hdr = 4

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=row_hdr, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border_thin

    # Dữ liệu mẫu thực tế
    sample_data = [
        [1, "PYC 1234/VT4", "Module đầu vào input", "IUX 760 MI dùng cho hệ thống DCS Foxboro", "Cái", 4, "Foxboro / Pháp", "3.82.63.134.ENG.00.000", 13559000, "=F5*I5", 0.1, "=J5*K5", "Vật tư thay thế tủ điều khiển DCS"],
        [2, "PYC 1234/VT4", "Mặt công tắc", "Dùng cho 2 thiết bị - Model 6802", "Cái", 5, "Sino / Việt Nam", "3.34.40.292.VIE.00.000", 45000, "=F6*I6", 0.1, "=J6*K6", "Vật tư điện chiếu sáng hạ thế"],
        [3, "PYC 1234/VT4", "Hạt công tắc", "Hạt công tắc 1 chiều 16A", "Cái", 10, "Sino / Việt Nam", "3.34.40.291.VIE.00.000", 18000, "=F7*I7", 0.1, "=J7*K7", "Thiết bị đóng cắt"],
    ]

    font_data = Font(name="Times New Roman", size=10)
    for row_idx, r_vals in enumerate(sample_data, start=5):
        ws.append(r_vals)
        for col_idx in range(1, len(headers) + 1):
            c = ws.cell(row=row_idx, column=col_idx)
            c.font = font_data
            c.border = border_thin
            if col_idx in (1, 2, 5):
                c.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx in (6, 9, 10, 12):
                c.alignment = Alignment(horizontal="right", vertical="center")
                c.number_format = '#,##0'
            elif col_idx == 11:
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.number_format = '0%'
            else:
                c.alignment = Alignment(horizontal="left", vertical="center")

    # Column Widths
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 24
    ws.column_dimensions['D'].width = 38
    ws.column_dimensions['E'].width = 8
    ws.column_dimensions['F'].width = 16
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 24
    ws.column_dimensions['I'].width = 18
    ws.column_dimensions['J'].width = 18
    ws.column_dimensions['K'].width = 10
    ws.column_dimensions['L'].width = 16
    ws.column_dimensions['M'].width = 28

    template_path = os.path.join(os.path.abspath(os.path.dirname(__file__)), "data", "Mau_Du_Toan_Tham_Dinh.xlsx")
    wb.save(template_path)
    return send_file(template_path, as_attachment=True, download_name="Mau_Du_Toan_Tham_Dinh.xlsx")


if __name__ == "__main__":
    print("=" * 70)
    print("ThamDinhDuToanApp - To Tham Dinh Du Toan NMND Vinh Tan 4")
    print("Dia chi: http://localhost:5555")
    print("=" * 70)
    app.run(host="0.0.0.0", port=5555, debug=False)
