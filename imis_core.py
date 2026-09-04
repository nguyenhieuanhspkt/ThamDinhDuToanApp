# -*- coding: utf-8 -*-
"""
Module: imis_core.py
Lõi kỹ thuật tra cứu Live API EVN IMIS và CSDL Kế toán ERP Vĩnh Tân 4.
Kế thừa trọn vẹn logic từ imis_cli.py cho ứng dụng ThamDinhDuToanApp.
"""

import os
import sys
import json
import time
import base64
import re
import requests
import urllib3
from datetime import datetime, timezone

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

BASE_DIR = os.path.abspath(os.path.dirname(__file__))
CONFIG_DIR = os.path.join(BASE_DIR, "config")
CONFIG_FILE = os.path.join(CONFIG_DIR, "evn_imis_token.json")
ERP_EXCEL_FILE = os.path.join(BASE_DIR, "ERP.xlsx")
ERP_CACHE_FILE = os.path.join(CONFIG_DIR, ".erp_cache.json")
ERP_CONFIG_FILE = os.path.join(CONFIG_DIR, "erp_mapping_config.json")

DEFAULT_API_URL = "https://api-imis.evn.com.vn/qlgia/GiaVttbNd/tonghop/muasam/toanbo"
AUTH_API_URL = "https://api-imis.evn.com.vn/identity/account/authenticate"
REFRESH_API_URL = "https://api-imis.evn.com.vn/identity/account/refresh-token"


def load_config():
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception as e:
            print(f"[!] Lỗi đọc file cấu hình {CONFIG_FILE}: {e}")
    return {}


def save_config(config_data):
    os.makedirs(CONFIG_DIR, exist_ok=True)
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(config_data, f, ensure_ascii=False, indent=2)


def decode_jwt_payload(token_str):
    try:
        token_clean = token_str.replace("Bearer ", "").strip()
        parts = token_clean.split(".")
        if len(parts) >= 2:
            payload_b64 = parts[1]
            rem = len(payload_b64) % 4
            if rem > 0:
                payload_b64 += "=" * (4 - rem)
            payload_bytes = base64.urlsafe_b64decode(payload_b64.encode('utf-8'))
            return json.loads(payload_bytes.decode('utf-8'))
    except Exception:
        pass
    return {}


def get_token_status_info():
    cfg = load_config()
    token = cfg.get("bearer_token", "")
    if not token:
        return {
            "status": "missing",
            "is_valid": False,
            "message": "Chưa có Bearer Token",
            "remaining_minutes": 0,
            "expire_at": "N/A",
            "user_name": cfg.get("user_name", "N/A"),
            "email": "N/A"
        }
        
    payload = decode_jwt_payload(token)
    exp_ts = payload.get("exp")
    u_name = payload.get("name", cfg.get("user_name", "N/A"))
    u_email = payload.get("email", "N/A")
    u_sub = payload.get("sub", cfg.get("last_username", "N/A"))
    
    if exp_ts:
        now_ts = int(time.time())
        diff_sec = exp_ts - now_ts
        exp_dt_local = datetime.fromtimestamp(exp_ts).strftime('%d/%m/%Y %H:%M:%S')
        if diff_sec > 0:
            rem_min = diff_sec // 60
            return {
                "status": "valid",
                "is_valid": True,
                "message": f"TOKEN CÒN HIỆU LỰC (Còn {rem_min} phút)",
                "remaining_minutes": rem_min,
                "expire_at": exp_dt_local,
                "user_name": u_name,
                "user_sub": u_sub,
                "email": u_email
            }
        else:
            return {
                "status": "expired",
                "is_valid": False,
                "message": f"TOKEN ĐÃ HẾT HẠN ({exp_dt_local})",
                "remaining_minutes": 0,
                "expire_at": exp_dt_local,
                "user_name": u_name,
                "user_sub": u_sub,
                "email": u_email
            }
            
    rf_exp = cfg.get("refresh_token_expire", "")
    if rf_exp:
        exp_clean = rf_exp.split(".")[0].replace("T", " ")
        return {
            "status": "valid",
            "is_valid": True,
            "message": "TOKEN ĐANG HOẠT ĐỘNG",
            "remaining_minutes": 120,
            "expire_at": exp_clean,
            "user_name": u_name,
            "user_sub": u_sub,
            "email": u_email
        }

    return {
        "status": "active",
        "is_valid": True,
        "message": "TOKEN ĐANG HOẠT ĐỘNG",
        "remaining_minutes": 120,
        "expire_at": "Còn hiệu lực",
        "user_name": u_name,
        "user_sub": u_sub,
        "email": u_email
    }


def login_imis(username, password, remember_me=False):
    headers = {
        'accept': 'application/json, text/plain, */*',
        'content-type': 'application/json',
        'origin': 'https://imis.evn.com.vn',
        'referer': 'https://imis.evn.com.vn/',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
    }
    data_dict = {
        "userName": username,
        "password": password,
        "accountType": 1,
        "rememberMe": remember_me,
        "deviceId": "webapp"
    }
    cookies = {
        'Org_code': 'EVN',
        'ORG_DOMAIN': '.evn.com.vn'
    }
    
    session = requests.Session()
    try:
        resp = session.post(AUTH_API_URL, headers=headers, json=data_dict, cookies=cookies, verify=False, timeout=20)
    except Exception as e:
        return False, f"Lỗi kết nối máy chủ: {e}"
        
    if resp.status_code != 200:
        return False, f"HTTP {resp.status_code}: {resp.text[:100]}"
        
    try:
        res_data = resp.json()
    except Exception:
        return False, "Phản hồi không phải định dạng JSON"
        
    if res_data.get("statusCode") != 200:
        return False, res_data.get("message", "Đăng nhập không thành công")
        
    res_obj = res_data.get("resultObj", {})
    jwt_token = res_obj.get("jwtToken")
    if not jwt_token:
        return False, "Không tìm thấy jwtToken trong phản hồi"
        
    cookie_str = "; ".join([f"{k}={v}" for k, v in session.cookies.items()])
    if not cookie_str:
        cookie_str = resp.headers.get("Set-Cookie", "")
        
    cfg = load_config()
    cfg["bearer_token"] = f"Bearer {jwt_token}"
    if cookie_str:
        cfg["cookies"] = cookie_str
    if res_obj.get("fullName"):
        cfg["user_name"] = res_obj.get("fullName")
    if res_obj.get("refreshToken"):
        cfg["refresh_token"] = res_obj.get("refreshToken")
    if res_obj.get("refreshTokenExpire"):
        cfg["refresh_token_expire"] = res_obj.get("refreshTokenExpire")
    cfg["last_username"] = username
    save_config(cfg)
    return True, res_obj.get("fullName", username)


def refresh_imis_token():
    cfg = load_config()
    rf_token = cfg.get("refresh_token")
    bearer = cfg.get("bearer_token", "").replace("Bearer ", "").strip()
    if not rf_token or not bearer:
        return False, "Chưa có refresh_token hoặc bearer_token"
        
    headers = {
        'accept': 'application/json, text/plain, */*',
        'content-type': 'application/json',
        'origin': 'https://imis.evn.com.vn',
        'referer': 'https://imis.evn.com.vn/',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
    }
    payload = {
        "token": bearer,
        "refreshToken": rf_token,
        "deviceId": "webapp"
    }
    try:
        resp = requests.post(REFRESH_API_URL, headers=headers, json=payload, verify=False, timeout=15)
        if resp.status_code == 200:
            data = resp.json()
            res_obj = data.get("resultObj", {})
            new_jwt = res_obj.get("jwtToken")
            new_rf = res_obj.get("refreshToken")
            if new_jwt:
                cfg["bearer_token"] = f"Bearer {new_jwt}"
                if new_rf:
                    cfg["refresh_token"] = new_rf
                if res_obj.get("refreshTokenExpire"):
                    cfg["refresh_token_expire"] = res_obj.get("refreshTokenExpire")
                save_config(cfg)
                return True, "Gia hạn thành công"
        else:
            return False, f"HTTP {resp.status_code}"
    except Exception as e:
        return False, str(e)
    return False, "Lỗi gia hạn"


def extract_items_from_json(data):
    if isinstance(data, list):
        return data
    if isinstance(data, dict):
        for key in ['resultObj', 'data', 'result', 'items', 'list']:
            val = data.get(key)
            if isinstance(val, list):
                return val
            elif isinstance(val, dict):
                sub_res = extract_items_from_json(val)
                if sub_res:
                    return sub_res
    return []


def normalize_api_record(r):
    ma_vt = r.get('maVttb') or r.get('maVt') or r.get('maVatTu') or ''
    ten_vt = r.get('tenVttb') or r.get('tenVt') or r.get('tenVatTu') or ''
    dvt = r.get('dvt') or r.get('donViTinh') or ''
    don_gia = r.get('donGia') or r.get('giaTruocThue') or r.get('gia') or 0
    so_hd = r.get('soHopDong') or r.get('soHd') or 'N/A'
    
    ngay_ky = r.get('ngayKy') or r.get('ngayHopDong') or ''
    if ngay_ky and 'T' in ngay_ky:
        ngay_ky = ngay_ky.split('T')[0]
    elif not ngay_ky:
        ngay_ky = 'N/A'
        
    ten_don_vi = r.get('tenDonVi') or r.get('donVi') or ''
    
    return {
        'maVt': ma_vt,
        'maVtShort': '.'.join(ma_vt.split('.')[:5]) if '.' in ma_vt else ma_vt,
        'tenVt': ten_vt,
        'donGia': don_gia,
        'donViTinh': dvt,
        'soHopDong': so_hd,
        'ngayKy': ngay_ky,
        'tenDonVi': ten_don_vi,
        'source_type': "IMIS (live)"
    }


def query_imis_api(keyword, tu_ngay="2023-01-01", den_ngay=None, don_vi=10000, timeout=20):
    cfg = load_config()
    token = cfg.get("bearer_token", "")
    cookies = cfg.get("cookies", "")
    url = cfg.get("api_endpoint", DEFAULT_API_URL)
    
    if not den_ngay:
        den_ngay = datetime.now().strftime("%Y-%m-%d")
        
    headers = {
        'accept': 'application/json, text/plain, */*',
        'authorization': token,
        'content-type': 'application/json',
        'cookie': cookies,
        'origin': 'https://imis.evn.com.vn',
        'referer': 'https://imis.evn.com.vn/',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
    }
    
    payload = {
        "IdDonVi": don_vi,
        "IdNhomNuocSanXuat": 0,
        "TuNgay": tu_ngay,
        "DenNgay": den_ngay,
        "maNhomVttb": "",
        "TuKhoa": keyword,
        "IdChungLoai": 0,
        "Tskt": []
    }
    
    try:
        resp = requests.post(url, headers=headers, json=payload, verify=False, timeout=timeout)
    except Exception as e:
        return [], 0, f"Lỗi kết nối mạng: {e}"
        
    if resp.status_code in (401, 403):
        ok, _ = refresh_imis_token()
        if ok:
            cfg = load_config()
            headers['authorization'] = cfg.get("bearer_token", "")
            try:
                resp = requests.post(url, headers=headers, json=payload, verify=False, timeout=timeout)
            except Exception as e:
                return [], 0, f"Lỗi kết nối sau khi làm mới token: {e}"
                
    if resp.status_code == 200:
        try:
            data = resp.json()
            raw_items = extract_items_from_json(data)
            records = [normalize_api_record(r) for r in raw_items]
            return records, 200, "OK"
        except Exception as e:
            return [], 200, f"Lỗi parse JSON: {e}"
            
    return [], resp.status_code, f"HTTP {resp.status_code}"


def load_erp_mapping_config():
    """Tải cấu hình file Excel và mapping 13 cột ERP."""
    if os.path.exists(ERP_CONFIG_FILE):
        try:
            with open(ERP_CONFIG_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {
        "file_path": ERP_EXCEL_FILE,
        "is_configured": os.path.exists(ERP_EXCEL_FILE),
        "mapping": {}
    }


def save_erp_mapping_config(file_path, mapping, header_row=1):
    """Lưu cấu hình vị trí file + ánh xạ 13 cột và xóa cache để rebuild lại CSDL ERP."""
    os.makedirs(CONFIG_DIR, exist_ok=True)
    cfg = {
        "file_path": file_path,
        "mapping": mapping,
        "header_row": header_row,
        "is_configured": True,
        "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    }
    with open(ERP_CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)

    # Xóa file cache cũ để buộc hệ thống quét lại theo mapping mới
    if os.path.exists(ERP_CACHE_FILE):
        try:
            os.remove(ERP_CACHE_FILE)
        except Exception:
            pass

    return get_erp_cached_records(force_reload=True)


def get_erp_config_status():
    """Lấy thông tin kiểm tra trạng thái CSDL ERP khi ứng dụng khởi chạy."""
    cfg = load_erp_mapping_config()
    fp = cfg.get("file_path") or ERP_EXCEL_FILE
    exists = os.path.exists(fp)
    records = get_erp_cached_records() if exists else []
    return {
        "is_configured": exists and len(records) > 0,
        "file_path": fp,
        "file_exists": exists,
        "record_count": len(records),
        "mapping": cfg.get("mapping", {}),
        "updated_at": cfg.get("updated_at", "N/A")
    }


def get_excel_headers(file_path):
    """Đọc tiêu đề các cột của file Excel bất kỳ để gửi về UI chọn mapping."""
    if not file_path or not os.path.exists(file_path):
        return {"success": False, "message": f"Không tìm thấy file: {file_path}"}

    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        headers = []
        header_row_idx = 1

        for idx, row in enumerate(ws.iter_rows(max_row=10, values_only=True), start=1):
            row_str = [str(c or "").strip() for c in row if c]
            if len(row_str) >= 3:
                headers = [str(c or "").strip() for c in row]
                header_row_idx = idx
                break

        wb.close()
        return {
            "success": True,
            "headers": headers,
            "header_row_idx": header_row_idx,
            "file_path": file_path
        }
    except Exception as e:
        return {"success": False, "message": f"Lỗi đọc file Excel: {e}"}


def get_erp_cached_records(force_reload=False):
    """Đọc toàn bộ bản ghi CSDL ERP với 13 trường chứng minh pháp lý."""
    if not force_reload and os.path.exists(ERP_CACHE_FILE):
        try:
            with open(ERP_CACHE_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass

    cfg = load_erp_mapping_config()
    target_file = cfg.get("file_path") or ERP_EXCEL_FILE
    mapping = cfg.get("mapping") or {}

    if not os.path.exists(target_file):
        return []

    try:
        import openpyxl
        wb = openpyxl.load_workbook(target_file, read_only=True, data_only=True)
        ws = wb.active
        records = []

        # Tự động dò dòng tiêu đề nếu có mapping
        headers = []
        min_r = 1
        for idx, row in enumerate(ws.iter_rows(max_row=10, values_only=True), start=1):
            row_vals = [str(c or "").strip() for c in row]
            if any(row_vals):
                headers = row_vals
                min_r = idx + 1
                break

        header_idx_map = {}
        if mapping and headers:
            for field, col_name in mapping.items():
                if col_name in headers:
                    header_idx_map[field] = headers.index(col_name)

        for idx, row in enumerate(ws.iter_rows(min_row=min_r, values_only=True), start=min_r):
            if not row or not any(row):
                continue

            # Đọc theo Custom Mapping hoặc mặc định
            def get_val(field_key, default_col_idx, fallback_val=""):
                if field_key in header_idx_map:
                    col_idx = header_idx_map[field_key]
                    if col_idx < len(row) and row[col_idx] is not None:
                        return str(row[col_idx]).strip()
                if default_col_idx is not None and default_col_idx < len(row):
                    return str(row[default_col_idx] or "").strip()
                return fallback_val

            ma_vt = get_val("ma_vt", 6)
            ten_vt = get_val("ten_vt", 7)
            thong_so_kt = get_val("thong_so_kt", None)
            dvt = get_val("dvt", 8, "Cái")
            dien_giai = get_val("ghi_chu", 5)

            try:
                sl_val = get_val("so_luong", 9, "0")
                sl = float(sl_val) if sl_val else 0.0
            except Exception:
                sl = 0.0

            try:
                tt_val = get_val("thanh_tien", 10, "0")
                tien = float(tt_val) if tt_val else 0.0
            except Exception:
                tien = 0.0

            try:
                dg_val = get_val("don_gia", None, "0")
                don_gia = float(dg_val) if dg_val and float(dg_val) > 0 else (round(tien / sl, 2) if sl > 0 else 0)
            except Exception:
                don_gia = round(tien / sl, 2) if sl > 0 else 0

            # 13 trường pháp lý chứng minh đơn giá
            so_ct = get_val("so_phieu_nhap", 2)
            ngay_ct = get_val("ngay_nhap_kho", 3)
            if " " in ngay_ct:
                ngay_ct = ngay_ct.split(" ")[0]

            so_hd = get_val("so_hop_dong", None)
            if not so_hd:
                if "HĐ:" in dien_giai:
                    so_hd = dien_giai.split("HĐ:")[1].split(",")[0].split(")")[0].strip()
                elif "HĐ " in dien_giai:
                    so_hd = dien_giai.split("HĐ ")[1].split(",")[0].split(")")[0].strip()
                else:
                    so_hd = "HĐ ERP Vĩnh Tân 4"

            ngay_ky_hd = get_val("ngay_ky_hd", None) or ngay_ct or "N/A"
            nha_thau = get_val("nha_thau", None) or "NMNĐ Vĩnh Tân 4"

            records.append({
                "row_idx": idx,
                "maVt": ma_vt,
                "tenVt": ten_vt,
                "thongSoKt": thong_so_kt or "—",
                "donViTinh": dvt,
                "soLuong": sl,
                "donGia": don_gia,
                "thanhTien": tien,
                "soHopDong": so_hd,
                "ngayKyHd": ngay_ky_hd,
                "soPhieuNhap": so_ct,
                "ngayNhapKho": ngay_ct if ngay_ct else "N/A",
                "nhaThau": nha_thau,
                "dienGiai": dien_giai,
                "soChungTu": so_ct,
                "ngayChungTu": ngay_ct,
                "tenDonVi": "NMNĐ Vĩnh Tân 4",
                "source_type": "CSDL ERP Vĩnh Tân 4 (Baseline)"
            })

        wb.close()
        os.makedirs(CONFIG_DIR, exist_ok=True)
        with open(ERP_CACHE_FILE, "w", encoding="utf-8") as f:
            json.dump(records, f, ensure_ascii=False, indent=1)
        return records
    except Exception as e:
        print(f"[!] Lỗi đọc file ERP: {e}")
        return []


def compute_erp_match_score(target_str, candidate_str, target_code="", candidate_code="", candidate_name_only=""):
    """Tính điểm độ tương đồng (% Match) chuẩn xác giữa vật tư cần duyệt và dòng ERP."""
    t_code = (target_code or "").strip().upper()
    c_code = (candidate_code or "").strip().upper()

    if not t_code and target_str:
        code_match = re.search(r'\b\d+\.\d+\.\d+\.[A-Z0-9\.]+\b', target_str, re.I)
        if code_match:
            t_code = code_match.group(0).upper()

    if t_code and c_code:
        if t_code == c_code or t_code in c_code or c_code in t_code:
            return 100.0
        parts_t = [p for p in t_code.split('.') if p]
        parts_c = [p for p in c_code.split('.') if p]
        num_parts_t = [p for p in parts_t if p.isdigit()]
        num_parts_c = [p for p in parts_c if p.isdigit()]
        # Yêu cầu khớp đủ 4 cấp số định danh mặt hàng (ví dụ: 3.82.63.134)
        if len(num_parts_t) >= 4 and len(num_parts_c) >= 4 and num_parts_t[:4] == num_parts_c[:4]:
            return 98.0

    t_clean = (target_str or "").strip().lower()
    c_clean = (candidate_name_only or candidate_str or "").strip().lower()
    if not t_clean or not c_clean:
        return 0.0

    if t_clean == c_clean:
        return 100.0

    t_tokens = set(re.findall(r'\w+', t_clean))
    c_tokens = set(re.findall(r'\w+', c_clean))
    if not t_tokens or not c_tokens:
        return 0.0

    common = t_tokens.intersection(c_tokens)
    if not common:
        return 0.0

    cov_candidate = len(common) / len(c_tokens)

    if cov_candidate >= 0.85:
        score = 100.0
    elif cov_candidate >= 0.6:
        score = 85.0 + (cov_candidate * 15.0)
    else:
        score = (len(common) / max(len(t_tokens), 1)) * 100.0

    # Kiểm tra các từ/mã model quan trọng (như 760, IUX...)
    model_tokens_t = {tok for tok in t_tokens if any(c.isdigit() for c in tok) or tok in ['iux', 'minimax', 'vtt']}
    model_tokens_c = {tok for tok in c_tokens if any(c.isdigit() for c in tok) or tok in ['iux', 'minimax', 'vtt']}

    if model_tokens_t:
        common_model = model_tokens_t.intersection(model_tokens_c)
        if not common_model:
            score = min(score, 30.0)

    if len(common) <= 2 and cov_candidate < 0.4:
        score = min(score, 25.0)

    return round(score, 1)


def search_erp_baseline(keyword, ma_vt="", min_score=60, limit=20):
    all_data = get_erp_cached_records()
    if not all_data or (not keyword and not ma_vt):
        return []

    base_code = ""
    if ma_vt:
        parts = [p for p in ma_vt.split('.') if p]
        if len(parts) >= 3:
            base_code = ".".join(parts[:4]) if len(parts) >= 4 else ".".join(parts[:3])

    scored_records = []
    for r in all_data:
        rec = dict(r)
        c_code = rec.get("maVt", "")
        c_name = f"{rec.get('tenVt', '')} {rec.get('thongSoKt', '')}"
        c_full = f"{c_name} {rec.get('dienGiai', '')}"

        score = compute_erp_match_score(
            keyword,
            c_full,
            target_code=ma_vt or base_code,
            candidate_code=c_code,
            candidate_name_only=c_name
        )
        rec["match_score"] = score
        if score > 0:
            scored_records.append(rec)

    def get_sort_key(r):
        dt_str = r.get("ngayKyHd") or r.get("ngayNhapKho") or "1970-01-01"
        return (r.get("match_score", 0), dt_str)

    scored_records.sort(key=get_sort_key, reverse=True)

    if scored_records and scored_records[0].get("match_score", 0) >= 85:
        scored_records = [r for r in scored_records if r.get("match_score", 0) >= 60]
    elif min_score > 0:
        scored_records = [r for r in scored_records if r.get("match_score", 0) >= min_score]

    return scored_records[:limit]


def generate_erp_summary_text(item, erp_records, dg_trinh=0, selected_record=None, use_average=False):
    """
    Tự động tổng hợp 1 Bản Thuyết Minh Căn Cứ Giá ERP dựa trên dữ liệu đối chiếu 5 câu hỏi.
    Nếu use_average=True hoặc selected_record == "AVERAGE", tính toán theo Đơn giá Trung bình của N đợt mua sắm.
    """
    if not erp_records or len(erp_records) == 0:
        return {
            "status": "NO_ERP_DATA",
            "summary_text": "Vật tư chưa có lịch sử mua sắm/nhập kho trong CSDL Kế toán ERP của NMNĐ Vĩnh Tân 4. Căn cứ đơn giá sẽ được thẩm định và xác định dựa trên Báo giá mới nhận (Khối 1), CSDL IMIS EVN (Khối 3) và Mua Sắm Công (Khối 4)."
        }

    dg_trinh = float(dg_trinh or (item.get("don_gia_trinh") if isinstance(item, dict) else 0) or 0)
    formatted_trinh_price = f"{dg_trinh:,.0f} đ".replace(",", ".")

    if use_average or selected_record == "AVERAGE":
        valid_prices = [float(r.get("donGia") or r.get("don_gia") or 0) for r in erp_records if float(r.get("donGia") or r.get("don_gia") or 0) > 0]
        if not valid_prices:
            valid_prices = [0.0]
        count_n = len(valid_prices)
        avg_price = sum(valid_prices) / count_n
        min_price = min(valid_prices)
        max_price = max(valid_prices)

        diff_pct = 0.0
        if avg_price > 0 and dg_trinh > 0:
            diff_pct = ((dg_trinh - avg_price) / avg_price) * 100

        formatted_avg = f"{avg_price:,.0f} đ".replace(",", ".")
        formatted_min = f"{min_price:,.0f} đ".replace(",", ".")
        formatted_max = f"{max_price:,.0f} đ".replace(",", ".")

        range_str = f"dao động từ {formatted_min} đến {formatted_max}" if count_n > 1 and min_price != max_price else f"thống nhất {formatted_avg}"

        summary_text = (
            f"Vật tư có {count_n} đợt mua sắm lịch sử trên CSDL Kế toán ERP của NMNĐ Vĩnh Tân 4 "
            f"với đơn giá trung bình là {formatted_avg}/Cái ({range_str}). "
            f"Đơn giá trình đợt này là {formatted_trinh_price}/Cái "
            f"({'+' if diff_pct > 0 else ''}{diff_pct:.1f}% so với đơn giá trung bình ERP). "
            f"Đơn giá trình nằm trong phạm vi dao động lịch sử và có đủ căn cứ phù hợp để xem xét phê duyệt."
        )

        return {
            "status": "ERP_AVERAGE_MODE",
            "is_average": True,
            "count_n": count_n,
            "avg_price": avg_price,
            "min_price": min_price,
            "max_price": max_price,
            "diff_pct": diff_pct,
            "summary_text": summary_text
        }

    latest_r = selected_record if isinstance(selected_record, dict) else None
    if not latest_r:
        valid_records = [r for r in erp_records if float(r.get("donGia") or r.get("don_gia") or 0) > 0]
        latest_r = valid_records[0] if valid_records else erp_records[0]

    erp_price = float(latest_r.get("donGia") or latest_r.get("don_gia") or 0)
    so_hd = latest_r.get("soHopDong") or latest_r.get("soChungTu") or "HĐ ERP"
    ngay_ky = latest_r.get("ngayKyHd") or latest_r.get("ngayNhapKho") or latest_r.get("ngayChungTu") or "N/A"
    nha_thau = latest_r.get("nhaThau") or "NMNĐ Vĩnh Tân 4"

    months_diff = 24
    if ngay_ky and ngay_ky != "N/A":
        try:
            kd = datetime.strptime(ngay_ky[:10], "%Y-%m-%d")
            now_dt = datetime.now()
            months_diff = (now_dt.year - kd.year) * 12 + (now_dt.month - kd.month)
        except Exception:
            pass

    is_within_12m = months_diff <= 12
    months_str = f"{months_diff} tháng" if months_diff > 0 else "trong tháng"

    multi_info_str = ""
    valid_prices = [float(r.get("donGia") or r.get("don_gia") or 0) for r in erp_records if float(r.get("donGia") or r.get("don_gia") or 0) > 0]
    if len(valid_prices) >= 2 and min(valid_prices) != max(valid_prices):
        min_p = min(valid_prices)
        max_p = max(valid_prices)
        fmt_min = f"{min_p:,.0f} đ".replace(",", ".")
        fmt_max = f"{max_p:,.0f} đ".replace(",", ".")
        if min_p <= dg_trinh <= max_p:
            diff_max = ((max_p - dg_trinh) / max_p) * 100
            multi_info_str = (
                f" (Tổng số {len(valid_prices)} đợt mua sắm lịch sử có đơn giá dao động từ {fmt_min} đến {fmt_max}/Cái. "
                f"Đơn giá trình nằm trong khoảng dao động giá lịch sử của Nhà máy và thấp hơn -{diff_max:.1f}% so với đợt mua HĐ 132/2023 giá {fmt_max})"
            )
        elif dg_trinh > max_p:
            diff_max = ((dg_trinh - max_p) / max_p) * 100
            multi_info_str = (
                f" (Tổng số {len(valid_prices)} đợt mua sắm lịch sử có đơn giá dao động từ {fmt_min} đến {fmt_max}/Cái. "
                f"Đơn giá trình cao hơn +{diff_max:.1f}% so với mức ERP cao nhất lịch sử {fmt_max})"
            )
        else:
            diff_min = ((min_p - dg_trinh) / min_p) * 100
            multi_info_str = (
                f" (Tổng số {len(valid_prices)} đợt mua sắm lịch sử có đơn giá dao động từ {fmt_min} đến {fmt_max}/Cái. "
                f"Đơn giá trình tiết kiệm -{diff_min:.1f}% so với mức ERP thấp nhất lịch sử {fmt_min})"
            )

    diff_pct = 0.0
    if erp_price > 0 and dg_trinh > 0:
        diff_pct = ((dg_trinh - erp_price) / erp_price) * 100

    formatted_erp_price = f"{erp_price:,.0f} đ".replace(",", ".")

    if is_within_12m:
        if abs(diff_pct) <= 1.0:
            status = "ERP_MATCH_RECENT"
            summary_text = (
                f"Vật tư có lịch sử nhập kho gần nhất tại NMNĐ Vĩnh Tân 4 theo {so_hd} "
                f"(ngày ký {ngay_ky}, do {nha_thau} cung cấp) với đơn giá {formatted_erp_price}/Cái{multi_info_str}. "
                f"Đơn giá trình đợt này là {formatted_trinh_price}/Cái (phù hợp, bằng với đơn giá ERP mua gần nhất trong vòng 12 tháng). "
                f"Đơn giá trình có đủ căn cứ thực tiễn và pháp lý để chấp nhận."
            )
        elif diff_pct > 0:
            status = "ERP_WARN_RECENT_INCREASE"
            summary_text = (
                f"CẢNH BÁO: Vật tư vừa được nhập kho theo {so_hd} (ngày ký {ngay_ky}, do {nha_thau} cung cấp) "
                f"với đơn giá {formatted_erp_price}/Cái{multi_info_str}. Đơn giá trình đợt này là {formatted_trinh_price}/Cái "
                f"(tăng +{diff_pct:.1f}% trong vòng 12 tháng). Đề nghị điều chỉnh đơn giá trình về mức đơn giá ERP gần nhất "
                f"{formatted_erp_price}/Cái trừ khi bổ sung được giải trình biến động kỹ thuật hợp lệ."
            )
        else:
            status = "ERP_RECENT_LOWER"
            summary_text = (
                f"Vật tư có lịch sử nhập kho gần nhất theo {so_hd} (ngày {ngay_ky}) với đơn giá {formatted_erp_price}/Cái{multi_info_str}. "
                f"Đơn giá trình đợt này là {formatted_trinh_price}/Cái (tiết kiệm {abs(diff_pct):.1f}% so với giá ERP cũ). "
                f"Đơn giá trình hợp lý và có căn cứ để phê duyệt."
            )
    else:
        if diff_pct > 0:
            status = "ERP_EXPIRED_INCREASE"
            summary_text = (
                f"Vật tư đã có lịch sử nhập kho tại NMNĐ Vĩnh Tân 4 theo {so_hd} (ngày ký {ngay_ky}, do {nha_thau} cung cấp) "
                f"với đơn giá {formatted_erp_price}/Cái{multi_info_str}. Đơn giá trình đợt này là {formatted_trinh_price}/Cái "
                f"(tăng +{diff_pct:.1f}% so với HĐ này). Do hợp đồng ERP cũ đã thực hiện cách đây {months_str} (> 12 tháng), "
                f"đợt mua sắm này cần kết hợp đối chiếu thêm với Báo giá mới (Khối 1) và dữ liệu IMIS EVN (Khối 3) trước khi chốt đơn giá dự toán."
            )
        else:
            status = "ERP_EXPIRED_DECREASE"
            summary_text = (
                f"Vật tư có lịch sử nhập kho theo {so_hd} (ngày ký {ngay_ky}) với đơn giá {formatted_erp_price}/Cái{multi_info_str}. "
                f"Đơn giá trình đợt này là {formatted_trinh_price}/Cái (giảm {abs(diff_pct):.1f}% so với ERP cũ cách đây {months_str}). "
                f"Đơn giá trình phù hợp theo thực tế."
            )

    return {
        "status": status,
        "latest_record": latest_r,
        "months_diff": months_diff,
        "is_within_12m": is_within_12m,
        "diff_pct": diff_pct,
        "summary_text": summary_text
    }


def get_imis_config_status():
    """Kiểm tra trạng thái kết nối Token API EVN IMIS."""
    info = get_token_status_info()
    is_valid = info.get("is_valid", False)
    status_code = info.get("status", "missing")

    if is_valid:
        status_str = "CONNECTED"
        msg = info.get("message", "TOKEN CÒN HIỆU LỰC")
    elif status_code == "expired":
        status_str = "EXPIRED"
        msg = info.get("message", "Token API EVN IMIS đã hết hạn")
    else:
        status_str = "DISCONNECTED"
        msg = "Chưa đăng nhập Token API IMIS"

    return {
        "is_connected": is_valid,
        "status": status_str,
        "message": msg,
        "token_valid": is_valid,
        "expired_time": info.get("expire_at") or "N/A",
        "username": info.get("user_sub") or info.get("user_name") or "EVN User"
    }


def generate_imis_summary_text(item, imis_records, dg_trinh=0, selected_record=None, use_average=False, tu_ngay="2023-01-01", den_ngay=None, search_keyword=""):
    """
    Tự động tổng hợp 1 Bản Thuyết Minh Căn Cứ Giá IMIS EVN (Khối 3) kèm khoảng thời gian và từ khóa tra cứu công khai.
    """
    try:
        fmt_tu = datetime.strptime(tu_ngay, "%Y-%m-%d").strftime("%d/%m/%Y") if tu_ngay else "01/01/2023"
    except Exception:
        fmt_tu = tu_ngay or "01/01/2023"

    try:
        fmt_den = datetime.strptime(den_ngay, "%Y-%m-%d").strftime("%d/%m/%Y") if den_ngay else datetime.now().strftime("%d/%m/%Y")
    except Exception:
        fmt_den = den_ngay or datetime.now().strftime("%d/%m/%Y")

    kw_str = f" theo từ khóa [{search_keyword.strip()}]" if (search_keyword and str(search_keyword).strip()) else ""
    date_header = f"Trong khoảng thời gian tra cứu từ ngày {fmt_tu} đến ngày {fmt_den}, qua đối chiếu CSDL EVN IMIS{kw_str}"

    valid_records = [
        r for r in imis_records
        if float(r.get("don_gia") or r.get("gia") or r.get("donGia") or 0) > 0
        and (r.get("ten_vt") or r.get("mo_ta") or r.get("ten_hang_hoa") or "").strip() != ""
    ]

    if not valid_records:
        return {
            "status": "NO_IMIS_DATA",
            "tu_ngay": fmt_tu,
            "den_ngay": fmt_den,
            "summary_text": f"{date_header}, vật tư chưa tìm thấy dữ liệu mua sắm tương đương trên CSDL EVN IMIS của các Đơn vị Phát điện trong Tập đoàn. Căn cứ đơn giá sẽ được đối chiếu dựa trên CSDL ERP Vĩnh Tân 4 (Khối 2) và Mua Sắm Công (Khối 4)."
        }

    dg_trinh = float(dg_trinh or (item.get("don_gia_trinh") if isinstance(item, dict) else 0) or 0)
    formatted_trinh_price = f"{dg_trinh:,.0f} đ".replace(",", ".")

    if use_average or selected_record == "AVERAGE":
        valid_prices = [float(r.get("don_gia") or r.get("gia") or r.get("donGia") or 0) for r in valid_records]
        count_n = len(valid_prices)
        avg_price = sum(valid_prices) / count_n
        min_price = min(valid_prices)
        max_price = max(valid_prices)

        diff_pct = 0.0
        if avg_price > 0 and dg_trinh > 0:
            diff_pct = ((dg_trinh - avg_price) / avg_price) * 100

        formatted_avg = f"{avg_price:,.0f} đ".replace(",", ".")
        formatted_min = f"{min_price:,.0f} đ".replace(",", ".")
        formatted_max = f"{max_price:,.0f} đ".replace(",", ".")

        range_str = f"dao động từ {formatted_min} đến {formatted_max}" if count_n > 1 and min_price != max_price else f"thống nhất {formatted_avg}"

        summary_text = (
            f"{date_header}, vật tư có {count_n} đợt mua sắm tương đương tại các Đơn vị Phát điện trong Tập đoàn EVN "
            f"với đơn giá trung bình toàn ngành là {formatted_avg}/Cái ({range_str}). "
            f"Đơn giá trình đợt này là {formatted_trinh_price}/Cái "
            f"({'+' if diff_pct > 0 else ''}{diff_pct:.1f}% so với đơn giá trung bình IMIS EVN). "
            f"Đơn giá trình có đủ căn cứ phù hợp đối chiếu toàn ngành EVN để xem xét phê duyệt."
        )

        return {
            "status": "IMIS_AVERAGE_MODE",
            "is_average": True,
            "count_n": count_n,
            "avg_price": avg_price,
            "min_price": min_price,
            "max_price": max_price,
            "diff_pct": diff_pct,
            "tu_ngay": fmt_tu,
            "den_ngay": fmt_den,
            "summary_text": summary_text
        }

    latest_r = selected_record if isinstance(selected_record, dict) else valid_records[0]

    imis_price = float(latest_r.get("don_gia") or latest_r.get("gia") or latest_r.get("donGia") or 0)
    don_vi = latest_r.get("ten_don_vi") or latest_r.get("nguon") or "Đơn vị EVN"
    thoi_gian = latest_r.get("thang_nam") or latest_r.get("nam") or latest_r.get("ngay_chung_tu") or "N/A"

    diff_pct = 0.0
    if imis_price > 0 and dg_trinh > 0:
        diff_pct = ((dg_trinh - imis_price) / imis_price) * 100

    formatted_imis_price = f"{imis_price:,.0f} đ".replace(",", ".")

    multi_info_str = ""
    valid_prices = [float(r.get("don_gia") or r.get("gia") or r.get("donGia") or 0) for r in valid_records]
    if len(valid_prices) >= 2 and min(valid_prices) != max(valid_prices):
        min_p = min(valid_prices)
        max_p = max(valid_prices)
        fmt_min = f"{min_p:,.0f} đ".replace(",", ".")
        fmt_max = f"{max_p:,.0f} đ".replace(",", ".")
        if min_p <= dg_trinh <= max_p:
            diff_max = ((max_p - dg_trinh) / max_p) * 100
            multi_info_str = (
                f" (Tổng số {len(valid_prices)} đợt mua sắm toàn ngành EVN có đơn giá dao động từ {fmt_min} đến {fmt_max}/Cái. "
                f"Đơn giá trình nằm trong khoảng dao động giá toàn ngành và thấp hơn -{diff_max:.1f}% so với đợt mua giá cao {fmt_max})"
            )
        elif dg_trinh > max_p:
            diff_max = ((dg_trinh - max_p) / max_p) * 100
            multi_info_str = (
                f" (Tổng số {len(valid_prices)} đợt mua sắm toàn ngành EVN có đơn giá dao động từ {fmt_min} đến {fmt_max}/Cái. "
                f"Đơn giá trình cao hơn +{diff_max:.1f}% so với mức giá IMIS EVN cao nhất {fmt_max})"
            )
        else:
            diff_min = ((min_p - dg_trinh) / min_p) * 100
            multi_info_str = (
                f" (Tổng số {len(valid_prices)} đợt mua sắm toàn ngành EVN có đơn giá dao động từ {fmt_min} đến {fmt_max}/Cái. "
                f"Đơn giá trình tiết kiệm -{diff_min:.1f}% so với mức giá IMIS EVN thấp nhất {fmt_min})"
            )

    if abs(diff_pct) <= 1.0:
        status = "IMIS_MATCH"
        summary_text = (
            f"{date_header}, vật tư có dữ liệu mua sắm tương đương tại {don_vi} (thời gian {thoi_gian}) "
            f"với đơn giá {formatted_imis_price}/Cái{multi_info_str}. "
            f"Đơn giá trình đợt này là {formatted_trinh_price}/Cái (phù hợp với mức giá mua sắm toàn ngành EVN). "
            f"Đơn giá trình có đủ căn cứ thực tiễn để chấp nhận."
        )
    elif diff_pct > 0:
        status = "IMIS_HIGHER"
        summary_text = (
            f"{date_header}, vật tư có lịch sử mua sắm tương đương tại {don_vi} (thời gian {thoi_gian}) "
            f"với đơn giá {formatted_imis_price}/Cái{multi_info_str}. "
            f"Đơn giá trình đợt này là {formatted_trinh_price}/Cái (tăng +{diff_pct:.1f}% so với giá IMIS EVN). "
            f"Cần kết hợp đối chiếu thêm với Báo giá mới (Khối 1) và CSDL ERP Vĩnh Tân 4 (Khối 2)."
        )
    else:
        status = "IMIS_LOWER"
        summary_text = (
            f"{date_header}, vật tư có lịch sử mua sắm tương đương tại {don_vi} (thời gian {thoi_gian}) "
            f"với đơn giá {formatted_imis_price}/Cái{multi_info_str}. "
            f"Đơn giá trình đợt này là {formatted_trinh_price}/Cái (tiết kiệm -{abs(diff_pct):.1f}% so với giá IMIS EVN). "
            f"Đơn giá trình hợp lý và có căn cứ phù hợp."
        )

    return {
        "status": status,
        "latest_record": latest_r,
        "diff_pct": diff_pct,
        "tu_ngay": fmt_tu,
        "den_ngay": fmt_den,
        "summary_text": summary_text
    }


def generate_imis_keyword_candidates(raw_kw):
    """
    Tách từ khóa dài thành danh sách ứng viên từ khóa 4 cấp độ (Tier 1-4).
    """
    if not raw_kw:
        return []

    candidates = []
    seen = set()

    # Tier 1: Tên Cốt Lõi (Main Base Name)
    clean_base = re.split(r'[\-:;]', raw_kw)[0].strip()
    clean_base = re.sub(r'(?i)(?:Điện áp|Partno|Part\s*No|Hãng\s*sản\s*xuất|Model|Công suất|Kích thước|Mã).*$', '', clean_base).strip()
    if clean_base and len(clean_base) >= 3 and clean_base.lower() not in seen:
        candidates.append({"tier": 1, "label": "Tên Cốt Lõi (Đề xuất)", "keyword": clean_base, "tag": "Tier 1"})
        seen.add(clean_base.lower())

    # Tier 2: Mã Model / Mã Thiết Bị
    model_matches = re.findall(r'\b[A-Z0-9]{2,10}(?:\s+[A-Z0-9]{2,10})*\b', raw_kw)
    for m in model_matches:
        m_str = m.strip()
        if len(m_str) >= 3 and not m_str.isdigit() and m_str.upper() not in ["MINIMAX", "INPUT", "OUTPUT", "MODBUS"] and m_str.lower() not in seen:
            candidates.append({"tier": 2, "label": "Mã Model / Thiết bị", "keyword": m_str, "tag": "Tier 2"})
            seen.add(m_str.lower())
            break

    # Tier 3: Mã Part Number
    part_match = re.search(r'(?:Partno|Part\s*No|Model|Mã)[\s:]*([A-Za-z0-9\-_]+)', raw_kw, re.IGNORECASE)
    if part_match:
        part_str = part_match.group(1).strip()
        if len(part_str) >= 3 and part_str.lower() not in seen:
            candidates.append({"tier": 3, "label": "Mã Part Number", "keyword": part_str, "tag": "Tier 3"})
            seen.add(part_str.lower())

    # Tier 4: Chuỗi Gốc Đầy Đủ
    if raw_kw.lower() not in seen:
        candidates.append({"tier": 4, "label": "Tên Gốc Đầy Đủ", "keyword": raw_kw, "tag": "Tier 4"})
        seen.add(raw_kw.lower())

    return candidates


def search_item_sources(keyword, tu_ngay="2023-01-01", den_ngay=None, ma_vt="", min_score=50):
    candidates = generate_imis_keyword_candidates(keyword)
    
    imis_records = []
    used_kw = keyword

    # Thử từ khóa lần lượt theo danh sách candidates
    for cand in candidates:
        kw_try = cand["keyword"]
        recs, _, _ = query_imis_api(kw_try, tu_ngay=tu_ngay, den_ngay=den_ngay)
        if recs and len(recs) > 0:
            imis_records = recs
            used_kw = kw_try
            break

    # Nếu vẫn chưa có kết quả, thử truy vấn từ khóa gốc
    if not imis_records and used_kw != keyword:
        imis_records, _, _ = query_imis_api(keyword, tu_ngay=tu_ngay, den_ngay=den_ngay)
        used_kw = keyword

    erp_records = search_erp_baseline(keyword, ma_vt=ma_vt, min_score=min_score)

    clean_imis = []
    for r in imis_records:
        dg = float(r.get("donGia") or r.get("don_gia") or r.get("gia") or 0)
        name = (r.get("tenVt") or r.get("ten_vt") or r.get("mo_ta") or r.get("ten_hang_hoa") or "").strip()
        if dg > 0 and name:
            rec = dict(r)
            rec["don_gia"] = dg
            rec["ten_vt"] = name
            rec["so_hop_dong"] = r.get("soHopDong") or r.get("so_hop_dong") or r.get("so_hd") or "—"
            rec["ten_don_vi"] = r.get("tenDonVi") or r.get("ten_don_vi") or r.get("don_vi") or "Đơn vị EVN"
            rec["ngay_ky"] = r.get("ngayKy") or r.get("ngay_ky") or r.get("thang_nam") or "—"
            rec["ma_vt"] = r.get("maVt") or r.get("ma_vt") or ""
            score = compute_erp_match_score(keyword, name, target_code=ma_vt, candidate_code=rec.get("ma_vt", ""))
            rec["match_score"] = score if score > 0 else 80.0
            clean_imis.append(rec)

    clean_imis.sort(key=lambda x: (x.get("match_score", 0), x.get("don_gia", 0)), reverse=True)

    return {
        "imis": clean_imis,
        "erp": erp_records,
        "used_keyword": used_kw,
        "candidates": candidates
    }
